# -*- coding: utf-8 -*-
"""Case_data_engineering_test.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1z1zkixvyT888JHR_B8oxVh5dLpw4HqU1
"""

pip install aspose-cells

from urllib import request
import jpype     
import asposecells  
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.pivot.fields import Missing
from openpyxl import workbook, load_workbook
from datetime import datetime

# File request and download 
file_url = 'http://raw.githubusercontent.com/raizen-analytics/data-engineering-test/master/assets/vendas-combustiveis-m3.xls'
filename = 'vendas_combustiveis_m3.xls'

request.urlretrieve(file_url , filename )

# Convert xls to xlsx file and save a raw_layer
root_path = '/content/'
raw_layer = 'raw'
raw_file = f'{raw_layer}_{filename[:-4]}.xlsx'

jpype.startJVM() 
from asposecells.api import Workbook
workbook = Workbook(f'{root_path}{filename}')
workbook.save(f'{root_path}{raw_file}')
jpype.shutdownJVM()

# Extract pivot table 'Sales of oil derivative fuels by UF and product'
raw_file_path = f'{root_path}{raw_file}'

workbook = load_workbook(raw_file_path)
worksheet = workbook['Plan1']
pivot_table_name = 'Tabela dinâmica1'

pivot_table = [p for p in worksheet._pivots if p.name == pivot_table_name][0]

# Extract a dict of all cache fields and their respective values
fields_map = {}
for field in pivot_table.cache.cacheFields:
    if field.sharedItems.count > 0:
        fields_map[field.name] = [f.v for f in field.sharedItems._fields]

# Extract all rows from cache records. Each row is initially parsed as a dict
column_names = [field.name for field in pivot_table.cache.cacheFields]
rows = []
for record in pivot_table.cache.records.r:
    # If some field in the record in missing, we replace it by NaN
    record_values = [
        field.v if not isinstance(field, Missing) else np.nan for field in record._fields
    ]

    row_dict = {k: v for k, v in zip(column_names, record_values)}

    # Shared fields are mapped as an Index, so we replace the field index by its value
    for key in fields_map:
        row_dict[key] = fields_map[key][row_dict[key]]

    rows.append(row_dict)

# Create a pandas dataframe and save a silver layer 
df = pd.DataFrame.from_dict(rows)

silver_layer = 'silver'
silver_file_name = 'sales_of_oil_derivative_fuels_by_UF_and_product'

df.to_parquet(f'{root_path}{silver_layer}_{silver_file_name}.parquet')

# Some transformations
df.drop(columns=['REGIÃO', 'TOTAL'], inplace=True)

df = df.melt(id_vars=['COMBUSTÍVEL','ANO', 'ESTADO', 'UNIDADE'], var_name = 'MES', value_name='VOLUME')
df['MES'] = df['MES'].replace(['Fev', 'Abr', 'Mai', 'Ago', 'Set', 'Out', 'Dez'], ['Feb', 'Apr', 'May', 'Aug', 'Sep', 'Oct', 'Dec'])

# Create date column
df = df.astype({'ANO' : int})
df = df.astype({'ANO' : str})
df['CONCAT_DT'] = df['ANO'] + '-' + df['MES'] + '-01'
df['DATE'] = pd.to_datetime(df.CONCAT_DT)
df.drop(columns=['ANO', 'MES', 'CONCAT_DT'], inplace=True)

# Rename columns
df = df[['DATE', 'ESTADO', 'COMBUSTÍVEL', 'UNIDADE', 'VOLUME']]
df.rename(columns={'DATE':'year_month', 'ESTADO' : 'uf', 'COMBUSTÍVEL' : 'product', 'UNIDADE' : 'unity', 'VOLUME' : 'volume' }, inplace=True)

df['created_at'] = datetime.now()

gold_layer = 'gold'
gold_file_name = silver_file_name

df.to_parquet(f'{root_path}{gold_layer}_{gold_file_name}.parquet')

# Display just to visualization
display(df)



# Extract pivot table 'Sales of diesel by UF and type'
raw_file_path = f'{root_path}{raw_file}'

workbook_1 = load_workbook(raw_file_path)
worksheet_1 = workbook_1['Plan1']
pivot_table_name = 'Tabela dinâmica3'

pivot_table = [p for p in worksheet_1._pivots if p.name == pivot_table_name][0]

# Extract a dict of all cache fields and their respective values
fields_map = {}
for field in pivot_table.cache.cacheFields:
    if field.sharedItems.count > 0:
        fields_map[field.name] = [f.v for f in field.sharedItems._fields]

# Extract all rows from cache records. Each row is initially parsed as a dict
column_names = [field.name for field in pivot_table.cache.cacheFields]
rows_1 = []
for record in pivot_table.cache.records.r:
    # If some field in the record in missing, we replace it by NaN
    record_values = [
        field.v if not isinstance(field, Missing) else np.nan for field in record._fields
    ]

    row_dict = {k: v for k, v in zip(column_names, record_values)}

    # Shared fields are mapped as an Index, so we replace the field index by its value
    for key in fields_map:
        row_dict[key] = fields_map[key][row_dict[key]]

    rows_1.append(row_dict)

# Create a pandas dataframe and save at silver layer 
df1 = pd.DataFrame.from_dict(rows_1)

silver_layer = 'silver'
silver_file_name = 'sales_of_diesel_by_UF_and_type'

df1.to_parquet(f'{root_path}{silver_layer}_{silver_file_name}.parquet')

# Some transformations
df1.drop(columns=['REGIÃO', 'TOTAL'], inplace=True)

df1 = df1.melt(id_vars=['COMBUSTÍVEL','ANO', 'ESTADO', 'UNIDADE'], var_name = 'MES', value_name='VOLUME')
df1['MES'] = df1['MES'].replace(['Fev', 'Abr', 'Mai', 'Ago', 'Set', 'Out', 'Dez'], ['Feb', 'Apr', 'May', 'Aug', 'Sep', 'Oct', 'Dec'])

# Create date column
df1 = df1.astype({'ANO' : int})
df1 = df1.astype({'ANO' : str})
df1['CONCAT_DT'] = df1['ANO'] + '-' + df1['MES'] + '-01'
df1['DATE'] = pd.to_datetime(df1.CONCAT_DT)
df1.drop(columns=['ANO', 'MES', 'CONCAT_DT'], inplace=True)

# Rename columns
df1 = df1[['DATE', 'ESTADO', 'COMBUSTÍVEL', 'UNIDADE', 'VOLUME']]
df1.rename(columns={'DATE':'year_month', 'ESTADO' : 'uf', 'COMBUSTÍVEL' : 'product', 'UNIDADE' : 'unity', 'VOLUME' : 'volume' }, inplace=True)

df1['created_at'] = datetime.now()

gold_layer = 'gold'
gold_file_name = silver_file_name

df1.to_parquet(f'{root_path}{gold_layer}_{gold_file_name}.parquet')

# Display just to visualization
display(df1)